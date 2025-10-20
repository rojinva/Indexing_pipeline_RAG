import os
import io
import openpyxl
import pandas as pd
import hashlib
from urllib.parse import unquote
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from azure.identity import ClientSecretCredential
from azure.storage.blob import BlobServiceClient
from .text_splitter import split_text_into_chunks
from .constants import blob_uri_prefix_constant

def get_blob_service_client(tenant_id, client_id, client_secret, storage_account_name):
    account_url = f"https://{storage_account_name}.blob.core.windows.net"
    credential = ClientSecretCredential(tenant_id, client_id, client_secret)
    blob_service_client = BlobServiceClient(
        account_url=account_url, credential=credential
    )
    return blob_service_client

def get_blob_filebuffer_and_metadata(blob_uri: str):
    try:
        tenant_id = os.environ["TENET_ID"]
        client_id = os.environ["CLIENT_ID"]
        client_secret = os.environ["CLIENT_SECRET"]
        storage_account_name = os.environ["STORAGE_ACCOUNT_NAME"]
        container_name = "knowledge-mining"  # Consider making this configurable

        blob_service_client = get_blob_service_client(
            tenant_id=tenant_id,
            client_id=client_id,
            client_secret=client_secret,
            storage_account_name=storage_account_name,
        )

        unquoted_blob_uri = unquote(blob_uri)
        parent_path = unquoted_blob_uri.split(container_name)[1]
        parent_path_with_container_name = unquoted_blob_uri.split(blob_uri_prefix_constant)[1]
        use_case = parent_path.split("/")[1] if "/" in parent_path else None
        parent_file_name = unquoted_blob_uri.rsplit("/", 1)[-1]

        blob_client = blob_service_client.get_blob_client(container=container_name, blob=parent_path).download_blob()
        blob_data = blob_client.content_as_bytes()
        file_buffer = io.BytesIO(blob_data)

        return file_buffer, parent_path_with_container_name, use_case, parent_path, parent_file_name
    except Exception as e:
        print(f"Error fetching blob file buffer and metadata: {e}")
        return None, None, None, None, None

def get_blob_size(tenant_id, client_id, client_secret, storage_account_name, container_name, blob_name):
    # Get the BlobServiceClient
    blob_service_client = get_blob_service_client(tenant_id, client_id, client_secret, storage_account_name)
    
    # Get the BlobClient
    blob_client = blob_service_client.get_blob_client(container=container_name, blob=blob_name)
    
    # Get the blob properties
    blob_properties = blob_client.get_blob_properties()
    
    # Return the size of the blob in MB
    size_in_bytes = blob_properties.size
    size_in_mb = size_in_bytes / (1024 * 1024)
    return size_in_mb

def generate_hash(text):
    hash_object = hashlib.sha256(text.encode())
    hash_hex = hash_object.hexdigest()
    return hash_hex


def split_excel_into_chunks(
    usecase: str, blob_uri: str, chunk_size: int, chunk_overlap: int
):
    if usecase == "commonSpec":
        return split_commonspec_file_into_chunks(blob_uri, chunk_size, chunk_overlap)
    elif usecase == "customerSurvey" or usecase == "DigitalTransformation" or usecase == "servicedesk" or usecase=="iplm-faq" or usecase=="demo-data" or usecase=="modelling-report" or usecase=="reliability-report" or usecase=="semis2_cci" or usecase=="sabre3d_kpr" or usecase=="sabre3d_bdsite" or usecase=="sabre3d_epl":
        return split_customersurvey_file_into_chunks(
            blob_uri, chunk_size, chunk_overlap
        )
    elif usecase == "edms" or usecase == "demoadlsfolder3":
        chunks, messages =  split_edms_excels_into_chunks(blob_uri, chunk_size, chunk_overlap)
        return chunks, messages
    elif usecase == "corporate-accounting":
        return split_corporate_accounting_excels_into_chunks(blob_uri, chunk_size, chunk_overlap)
    else:
        raise ValueError("Invalid usecase: must be 'commonSpec' or 'customerSurvey' or 'edms' or 'DigitalTransformation'.")


def find_common_spec_sheets(wb):
    """
    Determine which sheets in a given workbook are "common spec sheets."
    Common spec sheets will be pre-processed as individual single-row-plus-header files.
    A sheet is considered a common spec sheet if it contains a "common spec header" within the first 10 rows.
    A "common spec header" is a row that contains at least 3 of the following matches (case-insensitive):
        1. section OR module OR file name
        2. requirements OR specification OR details
        3. response
        4. clarification OR remark OR comment

    Parameters
    ----------
    wb : openpyxl.workbook.workbook.Workbook
        Workbook to parse.

    Returns
    -------
    sheet_config : dict (str -> (Bool, int, openpyxl.worksheet.worksheet.Worksheet))
        Dictionary mapping sheetname to 3-tuple: (1) whether the sheet is a common spec sheet, (2) header row, (3) the worksheet.
    """

    keyword_lists = [
        ["section", "module", "file name"],
        ["requirement", "spec", "details"],
        ["response", "compliance", "result"],
        ["clarification", "remark", "comment"],
    ]
    sheet_config = {}

    # Parse each sheet for possible common spec headers
    for ws in wb:

        sheet_config[ws.title] = (
            False,
            None,
            ws,
        )  # By default, not a common spec sheet

        for i, row in enumerate(ws):

            # Extract CSV-like text for this row
            values = [cell.value for cell in row]
            values = [str(v) for v in values]
            text = ",".join(values)
            text = text.lower()

            # Check row for header keyword matches
            keyword_count = sum(
                [any([kw in text for kw in kw_list]) for kw_list in keyword_lists]
            )
            if keyword_count >= 3:
                sheet_config[ws.title] = (True, i, ws)
                print(f"Identified common spec sheet (header={i}): {ws.title}")
                break

            # Bail if header is not found in first 10 rows
            if i >= 9:
                break

    return sheet_config


def split_commonspec_file_into_chunks(blob_uri, chunk_size, chunk_overlap):
    file_buffer, parent_path_with_container_name, use_case,parent_path,_ = get_blob_filebuffer_and_metadata(blob_uri)

    chunks_with_metadata = []

    if parent_path.endswith(".xls"):
        all_sheets = pd.read_excel(file_buffer, header=None, sheet_name=None)
        wb = Workbook()
        wb.remove(wb.active)

        for sheet_name, df in all_sheets.items():
            ws = wb.create_sheet(title=sheet_name)

            for r in dataframe_to_rows(df, index=False, header=False):
                ws.append(r)

    elif parent_path.endswith(".xlsx"):
        wb = openpyxl.load_workbook(file_buffer, read_only=True, data_only=True)

    # Determine common spec sheets
    sheet_config = find_common_spec_sheets(wb)

    # Iterate over each sheet
    for sheet_name, (is_common_spec, header, ws) in sheet_config.items():
        data = ws.values  # Get all rows
        data = list(data)[header:]  # Skip non-header rows
        df = pd.DataFrame(data)  # Convert to DataFrame

        # For common spec sheets, save individual rows
        if is_common_spec:
            # Convert the sheet to a pd.DataFrame

            df.columns = df.iloc[0]  # Set the header row
            df = df.iloc[1:]  # Skip the header row
            df = df.reset_index(drop=True)  # Reset the index

            # Drop empty rows and columns
            df=df.dropna(how="all", axis=0)
            df=df.dropna(how="all", axis=1)

            # Save each row (with header) as new text
            for i, row in df.iterrows():

                # Compute corresponding row number in the original Excel sheet
                row_number = (
                    # Current non-header row
                    i
                    +
                    # Number of skipped rows (above the header)
                    header
                    +
                    # Skip the header
                    1
                    +
                    # Excel is 1-indexed
                    1
                )

                # Drop columns with missing data
                row.dropna(inplace=True)

                # Convert Series to single-row DataFrame, and extract CSV text
                single_row_df = row.to_frame().T
                text = single_row_df.to_csv(None, index=False, header=True)
                chunked_text = split_text_into_chunks(
                    content=text, chunk_size=chunk_size, chunk_overlap=chunk_overlap
                )

                for text in chunked_text:
                    data = {
                        "content": text,
                        "chunk_hash": generate_hash(text),
                        "use_case": use_case,
                        "sheet_name": sheet_name,
                        "row": row_number,
                        "parent_path": parent_path_with_container_name,
                    }
                    chunks_with_metadata.append(data)

        else:
            # Get sheet contents
            text = df.to_csv(None, index=False, header=True)
            chunked_text = split_text_into_chunks(
                content=text, chunk_size=chunk_size, chunk_overlap=chunk_overlap
            )

            for text in chunked_text:
                data = {
                    "content": text,
                    "chunk_hash": generate_hash(text),
                    "use_case": use_case,
                    "sheet_name": sheet_name,
                    "row": -1,
                    "parent_path": parent_path_with_container_name,
                }
                chunks_with_metadata.append(data)

    return chunks_with_metadata


def split_customersurvey_file_into_chunks(blob_uri, chunk_size, chunk_overlap):
    file_buffer, parent_path_with_container_name, use_case,_,_ = get_blob_filebuffer_and_metadata(blob_uri)
    chunks_with_metadata = []

    wb = openpyxl.load_workbook(file_buffer, read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        data = sheet.values
        cols = next(data)
        data = list(data)

        
        df = pd.DataFrame(data, columns=cols)

        # Drop empty rows
        df=df.dropna(how="all", axis=0)

        for i in range(0, len(df)):
            header_plus_row_df = df.iloc[[i]]
            text = header_plus_row_df.to_csv(None, index=False, header=True)
            chunked_text = split_text_into_chunks(
                content=text, chunk_size=chunk_size, chunk_overlap=chunk_overlap
            )

            for text in chunked_text:
                data = {
                    "content": text,
                    "chunk_hash": generate_hash(text),
                    "use_case": use_case,
                    "sheet_name": sheet_name,
                    "row": i + 2,
                    "parent_path": parent_path_with_container_name,
                }
                chunks_with_metadata.append(data)

    return chunks_with_metadata

def split_csv_file_into_chunks(blob_uri, chunk_size, chunk_overlap):
    file_buffer, parent_path_with_container_name, use_case,_,_ = get_blob_filebuffer_and_metadata(blob_uri)

    chunks_with_metadata = []

    # Read the CSV file into a DataFrame with a specified encoding
    try:
        df = pd.read_csv(file_buffer, encoding='utf-8')
    except UnicodeDecodeError:
        # Try a different encoding if utf-8 fails
        df = pd.read_csv(file_buffer, encoding='ISO-8859-1')

    # Drop empty rows
    df=df.dropna(how="all", axis=0)

    for i in range(0, len(df)):
        header_plus_row_df = df.iloc[[i]]
        text = header_plus_row_df.to_csv(None, index=False, header=True)
        chunked_text = split_text_into_chunks(
            content=text, chunk_size=chunk_size, chunk_overlap=chunk_overlap
        )

        for text in chunked_text:
            data = {
                "content": text,
                "chunk_hash": generate_hash(text),
                "use_case": use_case,
                "sheet_name": "",
                "row": i + 2,
                "parent_path": parent_path_with_container_name,
            }
            chunks_with_metadata.append(data)

    return chunks_with_metadata


############# Adding corporate_accounting specific functions for excel splitting ###############
def split_corporate_accounting_excels_into_chunks(blob_uri, chunk_size, chunk_overlap):
    """
    Split corporate_accounting Excel files into batches of 50-100 rows while preserving headers.

    Args:
        blob_uri: URI of the blob to process
        chunk_size: Maximum size of each chunk
        chunk_overlap: Overlap between chunks
        
    Returns:
        List of chunks with metadata
    """
    file_buffer, parent_path_with_container_name, use_case, _, parent_file_name = get_blob_filebuffer_and_metadata(blob_uri)
    chunks_with_metadata = []
    
    if not file_buffer:
        return chunks_with_metadata

    batch_size = 50
    
    try:
        # Load workbook
        wb = openpyxl.load_workbook(file_buffer, read_only=True, data_only=True)

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            try:
                # Get sheet data
                data = sheet.values
                cols = next(data)
                data = list(data)
                
                # Create DataFrame
                df = pd.DataFrame(data, columns=cols)
                
                # Drop empty rows
                df = df.dropna(how="all", axis=0)
                
                if df.empty:
                    continue
                
                # Process in batches of 50-100 rows
                total_rows = len(df)
                for batch_start in range(0, total_rows, batch_size):
                    batch_end = min(batch_start + batch_size, total_rows)
                    batch_df = df.iloc[batch_start:batch_end]
                    
                    # Create chatbot-friendly format with headers preserved
                    content_parts = []
                    content_parts.append(f"Document: {parent_file_name}")
                    content_parts.append(f"Sheet: {sheet_name}")
                    content_parts.append(f"Rows: {batch_start + 2} to {batch_end + 1}")  # +2 for Excel indexing and header
                    content_parts.append("=" * 50)
                    
                    # Add column information
                    valid_columns = [str(col) for col in df.columns if pd.notna(col) and str(col).strip()]
                    if valid_columns:
                        content_parts.append(f"Columns: {', '.join(valid_columns)}")
                        content_parts.append("-" * 30)
                    
                    # Process each row in the batch
                    for idx, (_, row) in enumerate(batch_df.iterrows()):
                        row_parts = []
                        row_parts.append(f"Row {batch_start + idx + 2}:")  # +2 for Excel indexing and header
                        
                        # Create structured key-value pairs
                        for col, value in zip(df.columns, row):
                            if pd.notna(col) and pd.notna(value):
                                clean_col = str(col).strip()
                                clean_value = str(value).strip()
                                if clean_col and clean_value and clean_value.lower() not in ['nan', 'none', '']:
                                    row_parts.append(f"  {clean_col}: {clean_value};")
                        
                        # Join row parts with spaces and add to content_parts
                        content_parts.append(" ".join(row_parts))
                    
                    # Join content and create chunks if needed
                    full_content = "\n".join(content_parts)
                    
                    if len(full_content) <= chunk_size:
                        # Content fits in one chunk
                        chunk_texts = [full_content]
                    else:
                        # Split into multiple chunks using text splitter
                        chunk_texts = split_text_into_chunks(
                            content=full_content,
                            chunk_size=chunk_size,
                            chunk_overlap=chunk_overlap
                        )
                    
                    # Create metadata for each chunk
                    for chunk_idx, chunk_text in enumerate(chunk_texts):
                        chunk_data = {
                            "content": chunk_text,
                            "chunk_hash": generate_hash(chunk_text),
                            "use_case": use_case,
                            "sheet_name": sheet_name,
                            "row": batch_start + 2,  # +2 for Excel indexing and header
                            "parent_path": parent_path_with_container_name,
                        }
                        chunks_with_metadata.append(chunk_data)
                        
            except Exception as e:
                print(f"Error processing sheet {sheet_name} in corporate accounting file: {e}")
                continue
                
    except Exception as e:
        print(f"Error processing corporate accounting Excel file: {e}")
        
    return chunks_with_metadata


############# Adding enggelt specific functions for excel splitting ###############
def split_enggelt_file_into_chunks(blob_uri, chunk_size, chunk_overlap):
    def find_string_in_sheet(df, txt_to_search, rows_to_iterate, cols_to_iterate=5):
        df = df.reset_index()
        row_index = None
        for idx, row in df.iloc[:rows_to_iterate, :cols_to_iterate].iterrows():
            if txt_to_search.lower() in [str(value).lower() for value in row.values]:
                row_index = idx
                break
        return row_index

    def get_min_or_not_none(a, b):
        if a is None:
            return b
        elif b is None:
            return a
        elif a == b:
            return None
        else:
            return min(a, b)

    def delete_hidden_sheets(wb):
        all_sheets = wb.sheetnames

        for sheet_name in all_sheets:
            sheet = wb[sheet_name]

            if sheet.sheet_state == "hidden":
                del wb[sheet_name]

        return wb
    
    file_buffer, parent_path_with_container_name, use_case,_,_ = get_blob_filebuffer_and_metadata(blob_uri)

    chunks_with_metadata = []
    exclude_sheets = [
        "Archived Reference Material",
        "Variables",
        "Subsequent Sheet, Portrait ",
        "Subsequent Sheet, Landscape",
        "Subsequent Sheet, Freeform",
    ]
    rows_to_iterate = {"revision": 5}

    wb = openpyxl.load_workbook(file_buffer, read_only=True)
    wb = delete_hidden_sheets(wb)

    for sheet_name in wb.sheetnames:
        if sheet_name in exclude_sheets:
            continue
        sheet = wb[sheet_name]
        try:  # this is needed since some sheets have chart objects which cannot be read by openpyxl
            data = sheet.values
        except Exception as e:
            print(f"Error: {str(e)}")
            continue
        try:
            cols = next(data)
        except StopIteration:
            continue
        data = list(data)
        
        df = pd.DataFrame(data, columns=cols)
        # process revision history sheets if applicable
        if "revision history" in sheet_name.lower():
            row_start = find_string_in_sheet(df, "Rev", rows_to_iterate["revision"])
            if row_start != None:
                df = df.iloc[row_start:]
                df.columns = df.iloc[0]
                df = df[1:]
                # Drop empty rows and columns
                df=df.dropna(how="all", axis=0)
                df=df.dropna(how="all", axis=1)
                df["filename"] = parent_file_name
                df["sheet_name"] = sheet_name

                for i in range(0, len(df)):
                    header_plus_row_df = df.iloc[[i]]
                    text = header_plus_row_df.to_csv(None, index=False, header=True)
                    chunked_text = split_text_into_chunks(
                        content=text,
                        chunk_size=chunk_size,
                        chunk_overlap=chunk_overlap,
                    )

                    for text in chunked_text:
                        data = {
                            "content": text,
                            "chunk_hash": generate_hash(text),
                            "use_case": use_case,
                            "sheet_name": sheet_name,
                            "row": i + row_start + 3,
                            "parent_path": parent_path_with_container_name,
                        }
                        chunks_with_metadata.append(data)
                continue
        if "cover sheet" in sheet_name.lower():
            df=df.dropna(how="all", axis=0)
            df=df.dropna(how="all", axis=1)
            text = df.to_csv(None, index=False, header=False)
            chunked_text = split_text_into_chunks(
                content=text, chunk_size=chunk_size, chunk_overlap=chunk_overlap
            )

            for text in chunked_text:
                data = {
                    "content": text,
                    "chunk_hash": generate_hash(text),
                    "use_case": use_case,
                    "sheet_name": sheet_name,
                    "row": -1,
                    "parent_path": parent_path_with_container_name,
                }
                chunks_with_metadata.append(data)
            continue

        # check for sheets containing abbreviations
        if "abbreviations" in sheet_name.lower():
            row_start_abbreviation = find_string_in_sheet(df, "abbreviation", 10)
            if row_start_abbreviation != None:
                df = df.iloc[row_start_abbreviation:]
                df.columns = df.iloc[0]
                df = df[1:]
                # Drop empty rows and columns
                df=df.dropna(how="all", axis=0)
                df=df.dropna(how="all", axis=1)
                df["filename"] = parent_file_name
                df["sheet_name"] = sheet_name
                for i in range(0, len(df)):
                    header_plus_row_df = df.iloc[[i]]
                    text = header_plus_row_df.to_csv(None, index=False, header=True)
                    chunked_text = split_text_into_chunks(
                        content=text,
                        chunk_size=chunk_size,
                        chunk_overlap=chunk_overlap,
                    )

                    for text in chunked_text:
                        data = {
                            "content": text,
                            "chunk_hash": generate_hash(text),
                            "use_case": use_case,
                            "sheet_name": sheet_name,
                            "row": i + row_start_abbreviation + 3,
                            "parent_path": parent_path_with_container_name,
                        }
                        chunks_with_metadata.append(data)
                continue

        # now we test for sheets having the term 'parameter' or 'parameters' in the first 4 rows and only in the first column
        row_start_parameter_column = find_string_in_sheet(df, "parameter", 5, 2)
        row_start_parameters_column = find_string_in_sheet(df, "parameters", 5, 2)
        row_start_commodity = get_min_or_not_none(
            row_start_parameter_column, row_start_parameters_column
        )
        if row_start_commodity != None:
            df = df.iloc[row_start_commodity:]
            df.columns = df.iloc[0]
            df = df[1:]
            # Drop empty rows and columns
            df=df.dropna(how="all", axis=0)
            df=df.dropna(how="all", axis=1)
            df["filename"] = parent_file_name
            df["sheet_name"] = sheet_name
            for i in range(0, len(df)):
                header_plus_row_df = df.iloc[[i]]
                text = header_plus_row_df.to_csv(None, index=False, header=True)
                chunked_text = split_text_into_chunks(
                    content=text, chunk_size=chunk_size, chunk_overlap=chunk_overlap
                )

                for text in chunked_text:
                    data = {
                        "content": text,
                        "chunk_hash": generate_hash(text),
                        "use_case": use_case,
                        "sheet_name": sheet_name,
                        "row": i + row_start_commodity + 3,
                        "parent_path": parent_path_with_container_name,
                    }
                    chunks_with_metadata.append(data)
            continue
        else:
            if ("Unnamed: 0" not in list(df.columns)) or (
                "Unnamed: 1" not in list(df.columns)
            ):
                for i in range(0, len(df)):
                    header_plus_row_df = df.iloc[[i]]
                    header_plus_row_df=header_plus_row_df.dropna(axis=0, how="all")
                    header_plus_row_df["filename"] = parent_file_name
                    header_plus_row_df["sheet_name"] = sheet_name
                    if len(header_plus_row_df) == 0:
                        continue
                    text = header_plus_row_df.to_csv(None, index=False, header=True)
                    chunked_text = split_text_into_chunks(
                        content=text,
                        chunk_size=chunk_size,
                        chunk_overlap=chunk_overlap,
                    )

                    for text in chunked_text:
                        data = {
                            "content": text,
                            "chunk_hash": generate_hash(text),
                            "use_case": use_case,
                            "sheet_name": sheet_name,
                            "row": i + 2,
                            "parent_path": parent_path_with_container_name,
                        }
                        chunks_with_metadata.append(data)
                continue
            else:
                df=df.dropna(how="all", axis=0)
                df=df.dropna(how="all", axis=1)
                text = df.to_csv(None, index=False, header=False)
                chunked_text = split_text_into_chunks(
                    content=text, chunk_size=chunk_size, chunk_overlap=chunk_overlap
                )

                for text in chunked_text:
                    data = {
                        "content": text,
                        "chunk_hash": generate_hash(text),
                        "use_case": use_case,
                        "sheet_name": sheet_name,
                        "row": -1,
                        "parent_path": parent_path_with_container_name,
                    }
                    chunks_with_metadata.append(data)
                continue
    return chunks_with_metadata


def split_edms_excels_into_chunks(blob_uri, chunk_size, chunk_overlap):
    
    def delete_hidden_sheets(wb):
        all_sheets = wb.sheetnames

        for sheet_name in all_sheets:
            sheet = wb[sheet_name]

            if sheet.sheet_state == "hidden":
                del wb[sheet_name]

        return wb
    
    file_buffer, parent_path_with_container_name, use_case,_,_ = get_blob_filebuffer_and_metadata(blob_uri)

    chunks_with_metadata = []
    table_end_marker = "TABLE END"
    try:
        # Load the workbook with data_only=True to read values only
        wb = openpyxl.load_workbook(file_buffer, data_only=True)
        
        # Delete hidden sheets
        wb = delete_hidden_sheets(wb)
        
        # Initialize a dictionary to hold the processed data for each sheet
        processed_data = {}
        
        # Iterate through each sheet in the Excel file
        for sheet_name in wb.sheetnames:
            print(f"Processing sheet: {sheet_name}")
            # Read the sheet into a DataFrame
            sheet = wb[sheet_name]
            
            # Skip chartsheets
            if sheet.sheet_state == 'hidden' or isinstance(sheet, openpyxl.chartsheet.Chartsheet):
                print(f"Skipping chartsheet: {sheet_name}")
                continue

            data = sheet.values
            try:
                columns = next(data)[0:]
            except StopIteration:
                print(f"Sheet {sheet_name} is empty")
                continue
            # there is a possiblity that the column names are not unique, then in that scenario, add a prefix of 1,2,... etc to the duplicate columns
            columns = [f"{col}_{columns[:i].count(col) + 1}" if columns.count(col) > 1 else col for i, col in enumerate(columns)]
            df = pd.DataFrame(data, columns=columns)
            # Drop empty columns and rows
            df=df.dropna(axis=0, how='all')
            df=df.dropna(axis=1, how='all')
            # Replace None values with a blank or a single white space
            df.fillna(' ', inplace=True)

            if sheet_name.startswith("TBL"):
                # Add document name and sheet name as columns
                df['Document Name'] = parent_file_name
                df['Sheet Name'] = sheet_name

                # Forward fill to handle merged cells
                df.ffill(inplace=True)

                # Convert the first column to strings
                df.iloc[:, 0] = df.iloc[:, 0].astype(str)

                # Find the index of the "TABLE END" row
                table_end_index = df[df.iloc[:, 0].str.strip().str.upper() == table_end_marker].index
                if not table_end_index.empty:
                    table_end_index = table_end_index[0]
                    # Split the DataFrame into two parts
                    df_before = df.iloc[:table_end_index]
                    df_after = df.iloc[table_end_index + 1:]
                    
                    # Process rows before "TABLE END" as usual
                    data_list = df_before.to_dict(orient='records')
                    
                    # Process rows after "TABLE END" as a single text block
                    df_after = df_after.applymap(lambda x: x.strip() if isinstance(x, str) else x)
                    text_block = df_after.to_string(index=False).strip()
                    
                    # Add both parts to the processed data
                    processed_data[sheet_name] = {
                        "data_list": data_list,
                        "text_block": text_block
                    }
                else:
                    # If "TABLE END" is not found, process the entire DataFrame as usual
                    data_list = df.to_dict(orient='records')
                    processed_data[sheet_name] = {
                        "data_list": data_list,
                        "text_block": ""
                    }
            else:
                # Strip whitespace from each cell in the DataFrame
                df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
                # Convert DataFrame to Markdown
                markdown_text = df.to_string(index=False).strip() #df.to_markdown(index=False).strip()
                # Add the processed data to the dictionary
                processed_data[sheet_name] = {
                    "data_list": [],
                    "text_block": markdown_text
                }

        # processing the data into chunks with metadata
        for sheet_name, sheet_data in processed_data.items():
            data_list = sheet_data.get("data_list", [])
            text_block = sheet_data.get("text_block", "")

            # Process each row in data_list
            for i, row in enumerate(data_list):
                row_str = "{\n" + ",\n".join(f"  {k}: {v}" for k, v in row.items()) + "\n}"
                data={
                    "content": row_str,
                    "chunk_hash": generate_hash(row_str),
                    "use_case": use_case,
                    "sheet_name": sheet_name,
                    "row": i + 2,
                    "parent_path": parent_path_with_container_name
                }
                chunks_with_metadata.append(data)
            
            # Process the text block
            if text_block:
                chunked_text = split_text_into_chunks(content=text_block, chunk_size=chunk_size, chunk_overlap=chunk_overlap)
                for text in chunked_text:
                    prepend_str = "Document Name: " + parent_file_name + ",\n  Sheet Name: " + sheet_name + ",\n  Content: "
                    data = {
                        "content": prepend_str + text,
                        "chunk_hash": generate_hash(text),
                        "use_case": use_case,
                        "sheet_name": sheet_name,
                        "row": -1,
                        "parent_path": parent_path_with_container_name,
                    }
                    chunks_with_metadata.append(data)

        return chunks_with_metadata, ""
    except Exception as e:
        print(f"Error processing Excel file: {e}")
        return None, str(e)