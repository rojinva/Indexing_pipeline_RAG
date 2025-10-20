import os
import io
import openpyxl
import pandas as pd
import hashlib
from urllib.parse import unquote
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from azure.identity import ClientSecretCredential
from azure.storage.blob import BlobServiceClient

from .text_splitter import split_text_into_chunks


def get_blob_service_client(tenant_id, client_id, client_secret, storage_account_name):
    account_url = f"https://{storage_account_name}.blob.core.windows.net"
    credential = ClientSecretCredential(tenant_id, client_id, client_secret)
    blob_service_client = BlobServiceClient(
        account_url=account_url, credential=credential
    )
    return blob_service_client


def generate_hash(text):
    hash_object = hashlib.sha256(text.encode())
    hash_hex = hash_object.hexdigest()
    return hash_hex


def split_excel_into_chunks(
    usecase: str, blob_uri: str, chunk_size: int, chunk_overlap: int
):
    if usecase == "commonSpec":
        return split_commonspec_file_into_chunks(blob_uri, chunk_size, chunk_overlap)
    elif usecase == "customerSurvey":
        return split_customersurvey_file_into_chunks(
            blob_uri, chunk_size, chunk_overlap
        )
    elif usecase == "engineeringElt":
        return split_enggelt_file_into_chunks(blob_uri, chunk_size, chunk_overlap)
    else:
        raise ValueError("Invalid usecase: must be 'commonSpec' or 'customerSurvey'")


def find_common_spec_sheets(wb):
    """
    Determine which sheets in a given workbook are "common spec sheets."
    Common spec sheets will be pre-processed as individual single-row-plus-header files.
    A sheet is considered a common spec sheet if it contains a "common spec header" within the first 10 rows.
    A "common spec header" is a row that contains at least 3 of the following matches (case-insensitive):
        1. section OR module OR file name
        2. requirements OR specification OR details
        3. response
        4. clarification OR remark OR comment

    Parameters
    ----------
    wb : openpyxl.workbook.workbook.Workbook
        Workbook to parse.

    Returns
    -------
    sheet_config : dict (str -> (Bool, int, openpyxl.worksheet.worksheet.Worksheet))
        Dictionary mapping sheetname to 3-tuple: (1) whether the sheet is a common spec sheet, (2) header row, (3) the worksheet.
    """

    keyword_lists = [
        ["section", "module", "file name"],
        ["requirement", "spec", "details"],
        ["response", "compliance", "result"],
        ["clarification", "remark", "comment"],
    ]
    sheet_config = {}

    # Parse each sheet for possible common spec headers
    for ws in wb:

        sheet_config[ws.title] = (
            False,
            None,
            ws,
        )  # By default, not a common spec sheet

        for i, row in enumerate(ws):

            # Extract CSV-like text for this row
            values = [cell.value for cell in row]
            values = [str(v) for v in values]
            text = ",".join(values)
            text = text.lower()

            # Check row for header keyword matches
            keyword_count = sum(
                [any([kw in text for kw in kw_list]) for kw_list in keyword_lists]
            )
            if keyword_count >= 3:
                sheet_config[ws.title] = (True, i, ws)
                print(f"Identified common spec sheet (header={i}): {ws.title}")
                break

            # Bail if header is not found in first 10 rows
            if i >= 9:
                break

    return sheet_config


def split_commonspec_file_into_chunks(blob_uri, chunk_size, chunk_overlap):
    tenant_id = os.environ["TENET_ID"]
    client_id = os.environ["CLIENT_ID"]
    client_secret = os.environ["CLIENT_SECRET"]
    storage_account_name = os.environ["STORAGE_ACCOUNT_NAME"]
    container_name = "knowledge-mining"

    blob_service_client = get_blob_service_client(
        tenant_id=tenant_id,
        client_id=client_id,
        client_secret=client_secret,
        storage_account_name=storage_account_name,
    )

    # example parent path
    # parent_path = "https://dls2lamuswdfabdev001.blob.core.windows.net/knowledge-mining/commonSpec/Single%20Product%20Spec/ASE/Software%20Spec/C3/Sabre3D/2018/20181002_k21%208%20inch%20LAM%20Plate_Response.doc"
    # This path contains %20 and the url needs to be cleaned
    unquoted_blob_uri = unquote(blob_uri)
    parent_path = unquoted_blob_uri.split(container_name)[1]
    parent_path_with_container_name = unquoted_blob_uri.split(
        ".blob.core.windows.net/"
    )[1]
    use_case = parent_path.split("/")[1]

    blob_client = blob_service_client.get_blob_client(
        container=container_name, blob=parent_path
    ).download_blob()
    blob_data = blob_client.content_as_bytes()
    file_buffer = io.BytesIO(blob_data)

    chunks_with_metadata = []

    if parent_path.endswith(".xls"):
        all_sheets = pd.read_excel(file_buffer, header=None, sheet_name=None)
        wb = Workbook()
        wb.remove(wb.active)

        for sheet_name, df in all_sheets.items():
            ws = wb.create_sheet(title=sheet_name)

            for r in dataframe_to_rows(df, index=False, header=False):
                ws.append(r)

    elif parent_path.endswith(".xlsx"):
        wb = openpyxl.load_workbook(file_buffer, read_only=True, data_only=True)

    # Determine common spec sheets
    sheet_config = find_common_spec_sheets(wb)

    # Iterate over each sheet
    for sheet_name, (is_common_spec, header, ws) in sheet_config.items():
        data = ws.values  # Get all rows
        is_hidden = ws.sheet_state == "hidden"
        data = list(data)[header:]  # Skip non-header rows
        df = pd.DataFrame(data)  # Convert to DataFrame

        # For common spec sheets, save individual rows
        if is_common_spec:
            # Convert the sheet to a pd.DataFrame

            df.columns = df.iloc[0]  # Set the header row
            df = df.iloc[1:]  # Skip the header row
            df = df.reset_index(drop=True)  # Reset the index

            # Drop empty rows and columns
            df.dropna(how="all", axis=0, inplace=True)
            df.dropna(how="all", axis=1, inplace=True)

            # Save each row (with header) as new text
            for i, row in df.iterrows():

                # Compute corresponding row number in the original Excel sheet
                row_number = (
                    # Current non-header row
                    i
                    +
                    # Number of skipped rows (above the header)
                    header
                    +
                    # Skip the header
                    1
                    +
                    # Excel is 1-indexed
                    1
                )

                # Drop columns with missing data
                row.dropna(inplace=True)

                # Convert Series to single-row DataFrame, and extract CSV text
                single_row_df = row.to_frame().T
                text = single_row_df.to_csv(None, index=False, header=True)
                chunked_text = split_text_into_chunks(
                    content=text, chunk_size=chunk_size, chunk_overlap=chunk_overlap
                )

                for text in chunked_text:
                    data = {
                        "content": text,
                        "chunk_hash": generate_hash(text),
                        "use_case": use_case,
                        "sheet_name": sheet_name,
                        "row": row_number,
                        "parent_path": parent_path_with_container_name,
                    }
                    chunks_with_metadata.append(data)

        else:
            # Get sheet contents
            text = df.to_csv(None, index=False, header=True)
            chunked_text = split_text_into_chunks(
                content=text, chunk_size=chunk_size, chunk_overlap=chunk_overlap
            )

            for text in chunked_text:
                data = {
                    "content": text,
                    "chunk_hash": generate_hash(text),
                    "use_case": use_case,
                    "sheet_name": sheet_name,
                    "row": -1,
                    "parent_path": parent_path_with_container_name,
                }
                chunks_with_metadata.append(data)

    return chunks_with_metadata


def split_customersurvey_file_into_chunks(blob_uri, chunk_size, chunk_overlap):
    tenant_id = os.environ["TENET_ID"]
    client_id = os.environ["CLIENT_ID"]
    client_secret = os.environ["CLIENT_SECRET"]
    storage_account_name = os.environ["STORAGE_ACCOUNT_NAME"]
    container_name = "knowledge-mining"

    blob_service_client = get_blob_service_client(
        tenant_id=tenant_id,
        client_id=client_id,
        client_secret=client_secret,
        storage_account_name=storage_account_name,
    )

    # example parent path
    # parent_path = "https://dls2lamuswdfabdev001.blob.core.windows.net/knowledge-mining/commonSpec/Single%20Product%20Spec/ASE/Software%20Spec/C3/Sabre3D/2018/20181002_k21%208%20inch%20LAM%20Plate_Response.doc"
    # This path contains %20 and the url needs to be cleaned
    unquoted_blob_uri = unquote(blob_uri)
    parent_path = unquoted_blob_uri.split(container_name)[1]
    parent_path_with_container_name = unquoted_blob_uri.split(
        ".blob.core.windows.net/"
    )[1]
    use_case = parent_path.split("/")[1]

    blob_client = blob_service_client.get_blob_client(
        container=container_name, blob=parent_path
    ).download_blob()
    blob_data = blob_client.content_as_bytes()
    file_buffer = io.BytesIO(blob_data)
    chunks_with_metadata = []

    wb = openpyxl.load_workbook(file_buffer, read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        data = sheet.values
        cols = next(data)
        data = list(data)

        is_hidden = sheet.sheet_state == "hidden"
        df = pd.DataFrame(data, columns=cols)

        # Drop empty rows
        df.dropna(how="all", axis=0, inplace=True)

        for i in range(0, len(df)):
            header_plus_row_df = df.iloc[[i]]
            text = header_plus_row_df.to_csv(None, index=False, header=True)
            chunked_text = split_text_into_chunks(
                content=text, chunk_size=chunk_size, chunk_overlap=chunk_overlap
            )

            for text in chunked_text:
                data = {
                    "content": text,
                    "chunk_hash": generate_hash(text),
                    "use_case": use_case,
                    "sheet_name": sheet_name,
                    "row": i + 2,
                    "parent_path": parent_path_with_container_name,
                }
                chunks_with_metadata.append(data)

    return chunks_with_metadata


############# Adding enggelt specific functions for excel splitting ###############
def split_enggelt_file_into_chunks(blob_uri, chunk_size, chunk_overlap):
    def find_string_in_sheet(df, txt_to_search, rows_to_iterate, cols_to_iterate=5):
        df = df.reset_index()
        row_index = None
        for idx, row in df.iloc[:rows_to_iterate, :cols_to_iterate].iterrows():
            if txt_to_search.lower() in [str(value).lower() for value in row.values]:
                row_index = idx
                break
        return row_index

    def get_min_or_not_none(a, b):
        if a is None:
            return b
        elif b is None:
            return a
        elif a == b:
            return None
        else:
            return min(a, b)

    def delete_hidden_sheets(wb):
        all_sheets = wb.sheetnames

        for sheet_name in all_sheets:
            sheet = wb[sheet_name]

            if sheet.sheet_state == "hidden":
                del wb[sheet_name]

        return wb

    tenant_id = os.environ["TENET_ID"]
    client_id = os.environ["CLIENT_ID"]
    client_secret = os.environ["CLIENT_SECRET"]
    storage_account_name = os.environ["STORAGE_ACCOUNT_NAME"]
    container_name = "knowledge-mining"

    blob_service_client = get_blob_service_client(
        tenant_id=tenant_id,
        client_id=client_id,
        client_secret=client_secret,
        storage_account_name=storage_account_name,
    )

    # example parent path
    # parent_path = "https://dls2lamuswdfabdev001.blob.core.windows.net/knowledge-mining/commonSpec/Single%20Product%20Spec/ASE/Software%20Spec/C3/Sabre3D/2018/20181002_k21%208%20inch%20LAM%20Plate_Response.doc"
    # This path contains %20 and the url needs to be cleaned
    unquoted_blob_uri = unquote(blob_uri)
    parent_file_name = unquoted_blob_uri.rsplit("/", 1)[-1]
    parent_path = unquoted_blob_uri.split(container_name)[1]
    parent_path_with_container_name = unquoted_blob_uri.split(
        ".blob.core.windows.net/"
    )[1]
    use_case = parent_path.split("/")[1]

    blob_client = blob_service_client.get_blob_client(
        container=container_name, blob=parent_path
    ).download_blob()
    blob_data = blob_client.content_as_bytes()
    file_buffer = io.BytesIO(blob_data)
    chunks_with_metadata = []
    exclude_sheets = [
        "Archived Reference Material",
        "Variables",
        "Subsequent Sheet, Portrait ",
        "Subsequent Sheet, Landscape",
        "Subsequent Sheet, Freeform",
    ]
    rows_to_iterate = {"revision": 5}

    wb = openpyxl.load_workbook(file_buffer, read_only=True)
    wb = delete_hidden_sheets(wb)

    for sheet_name in wb.sheetnames:
        if sheet_name in exclude_sheets:
            continue
        sheet = wb[sheet_name]
        try:  # this is needed since some sheets have chart objects which cannot be read by openpyxl
            data = sheet.values
        except Exception as e:
            print(f"Error: {str(e)}")
            continue
        try:
            cols = next(data)
        except StopIteration:
            continue
        data = list(data)
        is_hidden = sheet.sheet_state == "hidden"
        df = pd.DataFrame(data, columns=cols)
        # process revision history sheets if applicable
        if "revision history" in sheet_name.lower():
            row_start = find_string_in_sheet(df, "Rev", rows_to_iterate["revision"])
            if row_start != None:
                df = df.iloc[row_start:]
                df.columns = df.iloc[0]
                df = df[1:]
                # Drop empty rows and columns
                df.dropna(how="all", axis=0, inplace=True)
                df.dropna(how="all", axis=1, inplace=True)
                df["filename"] = parent_file_name
                df["sheet_name"] = sheet_name

                for i in range(0, len(df)):
                    header_plus_row_df = df.iloc[[i]]
                    text = header_plus_row_df.to_csv(None, index=False, header=True)
                    chunked_text = split_text_into_chunks(
                        content=text,
                        chunk_size=chunk_size,
                        chunk_overlap=chunk_overlap,
                    )

                    for text in chunked_text:
                        data = {
                            "content": text,
                            "chunk_hash": generate_hash(text),
                            "use_case": use_case,
                            "sheet_name": sheet_name,
                            "row": i + row_start + 3,
                            "parent_path": parent_path_with_container_name,
                        }
                        chunks_with_metadata.append(data)
                continue
        if "cover sheet" in sheet_name.lower():
            df.dropna(how="all", axis=0, inplace=True)
            df.dropna(how="all", axis=1, inplace=True)
            text = df.to_csv(None, index=False, header=False)
            chunked_text = split_text_into_chunks(
                content=text, chunk_size=chunk_size, chunk_overlap=chunk_overlap
            )

            for text in chunked_text:
                data = {
                    "content": text,
                    "chunk_hash": generate_hash(text),
                    "use_case": use_case,
                    "sheet_name": sheet_name,
                    "row": -1,
                    "parent_path": parent_path_with_container_name,
                }
                chunks_with_metadata.append(data)
            continue

        # check for sheets containing abbreviations
        if "abbreviations" in sheet_name.lower():
            row_start_abbreviation = find_string_in_sheet(df, "abbreviation", 10)
            if row_start_abbreviation != None:
                df = df.iloc[row_start_abbreviation:]
                df.columns = df.iloc[0]
                df = df[1:]
                # Drop empty rows and columns
                df.dropna(how="all", axis=0, inplace=True)
                df.dropna(how="all", axis=1, inplace=True)
                df["filename"] = parent_file_name
                df["sheet_name"] = sheet_name
                for i in range(0, len(df)):
                    header_plus_row_df = df.iloc[[i]]
                    text = header_plus_row_df.to_csv(None, index=False, header=True)
                    chunked_text = split_text_into_chunks(
                        content=text,
                        chunk_size=chunk_size,
                        chunk_overlap=chunk_overlap,
                    )

                    for text in chunked_text:
                        data = {
                            "content": text,
                            "chunk_hash": generate_hash(text),
                            "use_case": use_case,
                            "sheet_name": sheet_name,
                            "row": i + row_start_abbreviation + 3,
                            "parent_path": parent_path_with_container_name,
                        }
                        chunks_with_metadata.append(data)
                continue

        # now we test for sheets having the term 'parameter' or 'parameters' in the first 4 rows and only in the first column
        row_start_parameter_column = find_string_in_sheet(df, "parameter", 5, 2)
        row_start_parameters_column = find_string_in_sheet(df, "parameters", 5, 2)
        row_start_commodity = get_min_or_not_none(
            row_start_parameter_column, row_start_parameters_column
        )
        if row_start_commodity != None:
            df = df.iloc[row_start_commodity:]
            df.columns = df.iloc[0]
            df = df[1:]
            # Drop empty rows and columns
            df.dropna(how="all", axis=0, inplace=True)
            df.dropna(how="all", axis=1, inplace=True)
            df["filename"] = parent_file_name
            df["sheet_name"] = sheet_name
            for i in range(0, len(df)):
                header_plus_row_df = df.iloc[[i]]
                text = header_plus_row_df.to_csv(None, index=False, header=True)
                chunked_text = split_text_into_chunks(
                    content=text, chunk_size=chunk_size, chunk_overlap=chunk_overlap
                )

                for text in chunked_text:
                    data = {
                        "content": text,
                        "chunk_hash": generate_hash(text),
                        "use_case": use_case,
                        "sheet_name": sheet_name,
                        "row": i + row_start_commodity + 3,
                        "parent_path": parent_path_with_container_name,
                    }
                    chunks_with_metadata.append(data)
            continue
        else:
            if ("Unnamed: 0" not in list(df.columns)) or (
                "Unnamed: 1" not in list(df.columns)
            ):
                for i in range(0, len(df)):
                    header_plus_row_df = df.iloc[[i]]
                    header_plus_row_df.dropna(axis=0, how="all", inplace=True)
                    header_plus_row_df["filename"] = parent_file_name
                    header_plus_row_df["sheet_name"] = sheet_name
                    if len(header_plus_row_df) == 0:
                        continue
                    text = header_plus_row_df.to_csv(None, index=False, header=True)
                    chunked_text = split_text_into_chunks(
                        content=text,
                        chunk_size=chunk_size,
                        chunk_overlap=chunk_overlap,
                    )

                    for text in chunked_text:
                        data = {
                            "content": text,
                            "chunk_hash": generate_hash(text),
                            "use_case": use_case,
                            "sheet_name": sheet_name,
                            "row": i + 2,
                            "parent_path": parent_path_with_container_name,
                        }
                        chunks_with_metadata.append(data)
                continue
            else:
                df.dropna(how="all", axis=0, inplace=True)
                df.dropna(how="all", axis=1, inplace=True)
                text = df.to_csv(None, index=False, header=False)
                chunked_text = split_text_into_chunks(
                    content=text, chunk_size=chunk_size, chunk_overlap=chunk_overlap
                )

                for text in chunked_text:
                    data = {
                        "content": text,
                        "chunk_hash": generate_hash(text),
                        "use_case": use_case,
                        "sheet_name": sheet_name,
                        "row": -1,
                        "parent_path": parent_path_with_container_name,
                    }
                    chunks_with_metadata.append(data)
                continue
    return chunks_with_metadata
